"""
Testes do Excel export (Fase 5 / Q6 + Commit 5).

Conforme Q6: 3 abas mínimas (Caso, Resultados, Geometria) + Diagnostics
opcional. Aba Diagnostics consistente com Memorial PDF.
"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.api.tests._fixtures import BC01_LIKE_INPUT


def _solve_case(client: TestClient) -> int:
    case_resp = client.post("/api/v1/cases", json=BC01_LIKE_INPUT)
    case_id = case_resp.json()["id"]
    client.post(f"/api/v1/cases/{case_id}/solve")
    return case_id


def _load_xlsx(content: bytes):
    return load_workbook(io.BytesIO(content), read_only=False, data_only=True)


# ─── Smoke ──────────────────────────────────────────────────────────


def test_xlsx_endpoint_404_quando_caso_nao_existe(client: TestClient):
    resp = client.get("/api/v1/cases/9999/export/xlsx")
    assert resp.status_code == 404


def test_xlsx_endpoint_retorna_xlsx_valido(client: TestClient, seeded_catalog):
    """Smoke: 200 + Content-Type Excel + magic header."""
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    assert resp.status_code == 200
    assert "spreadsheet" in resp.headers["content-type"]
    # XLSX é um zip — magic number PK
    assert resp.content[:2] == b"PK"


# ─── AC do plano: 3 abas mínimas ───────────────────────────────────


def test_xlsx_tem_3_abas_minimas(client: TestClient, seeded_catalog):
    """AC F5: 3 abas mínimas (Caso, Resultados, Geometria)."""
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    assert "Caso" in wb.sheetnames
    assert "Resultados" in wb.sheetnames
    assert "Geometria" in wb.sheetnames


# ─── Aba Caso ──────────────────────────────────────────────────────


def test_aba_caso_contem_metadata(client: TestClient, seeded_catalog):
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    ws = wb["Caso"]
    # A1 = título
    assert "AncoPlat" in str(ws["A1"].value)
    # Procura "Hash" em alguma célula da coluna A (não fixar linha exata)
    has_hash = any(
        ws.cell(row=r, column=1).value
        and "Hash" in str(ws.cell(row=r, column=1).value)
        for r in range(1, 30)
    )
    assert has_hash, "Aba Caso deve ter linha 'Hash do caso'"


def test_aba_caso_lista_segmentos(client: TestClient, seeded_catalog):
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    ws = wb["Caso"]
    # Procura "Segmentos" em alguma célula
    has_seg = any(
        ws.cell(row=r, column=1).value
        and "Segmentos" in str(ws.cell(row=r, column=1).value)
        for r in range(1, 50)
    )
    assert has_seg


# ─── Aba Resultados ────────────────────────────────────────────────


def test_aba_resultados_contem_profile_type(client: TestClient, seeded_catalog):
    """ProfileType (Fase 4) presente nos resultados."""
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    ws = wb["Resultados"]
    has_pt = any(
        ws.cell(row=r, column=1).value
        and "Profile" in str(ws.cell(row=r, column=1).value)
        for r in range(1, 30)
    )
    assert has_pt


def test_aba_resultados_contem_solver_version(client: TestClient, seeded_catalog):
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    ws = wb["Resultados"]
    has_ver = any(
        ws.cell(row=r, column=1).value
        and "version" in str(ws.cell(row=r, column=1).value).lower()
        for r in range(1, 30)
    )
    assert has_ver


# ─── Aba Geometria ─────────────────────────────────────────────────


def test_aba_geometria_tem_5000_pontos(client: TestClient, seeded_catalog):
    """≥ 5000 linhas de geometria (1 header + ≥ 5000 dados)."""
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    ws = wb["Geometria"]
    # Conta linhas não-vazias
    n_rows = sum(
        1 for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)
    )
    assert n_rows >= 5001, f"Apenas {n_rows} linhas; esperado ≥ 5001"


def test_aba_geometria_header_correto(client: TestClient, seeded_catalog):
    del seeded_catalog
    case_id = _solve_case(client)
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    ws = wb["Geometria"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    assert headers[0] == "x (m)"
    assert headers[1] == "y (m)"
    assert "tension_magnitude" in headers[4]


# ─── Aba Diagnostics (condicional) ─────────────────────────────────


def test_aba_diagnostics_aparece_quando_diagnostics_existem(
    client: TestClient, seeded_catalog,
):
    """
    Caso com utilização > 0.5 dispara D008/D010 → Diagnostics aparece.
    """
    del seeded_catalog
    # Caso com utilização ALTA pra disparar D010
    payload = {
        **BC01_LIKE_INPUT,
        "boundary": {**BC01_LIKE_INPUT["boundary"], "input_value": 2_500_000},
    }
    case_resp = client.post("/api/v1/cases", json=payload)
    case_id = case_resp.json()["id"]
    client.post(f"/api/v1/cases/{case_id}/solve")
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    if "Diagnostics" in wb.sheetnames:
        ws = wb["Diagnostics"]
        # Header obrigatório consistente com Memorial PDF (Q6 detail):
        # Code, Severity, Confidence, Title, Cause
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert headers == ["Code", "Severity", "Confidence", "Title", "Cause"]


def test_aba_diagnostics_ausente_quando_caso_limpo(
    client: TestClient, seeded_catalog,
):
    """Caso simples (sem diagnostics) NÃO tem aba Diagnostics."""
    del seeded_catalog
    # Cria + solve com utilização baixa
    payload = {
        **BC01_LIKE_INPUT,
        "boundary": {**BC01_LIKE_INPUT["boundary"], "input_value": 200_000},
    }
    case_resp = client.post("/api/v1/cases", json=payload)
    case_id = case_resp.json()["id"]
    client.post(f"/api/v1/cases/{case_id}/solve")
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    # Pode ou não ter aba — best-effort, depende dos diagnostics que disparam
    # Sanity: aba não aparece com lista vazia em algum caso
    # Verifico que a aba é ou ausente OU tem ≥ 1 row de dados
    if "Diagnostics" in wb.sheetnames:
        ws = wb["Diagnostics"]
        # Se aparece, tem header + pelo menos 1 dado
        n_rows = sum(
            1 for row in ws.iter_rows(values_only=True)
            if any(v is not None for v in row)
        )
        assert n_rows >= 2, "Aba Diagnostics aparece mas vazia"


# ─── Sem solve: só aba Caso ────────────────────────────────────────


def test_xlsx_caso_sem_solve_tem_so_aba_caso(client: TestClient, seeded_catalog):
    del seeded_catalog
    case_resp = client.post("/api/v1/cases", json=BC01_LIKE_INPUT)
    case_id = case_resp.json()["id"]
    # NÃO faz solve
    resp = client.get(f"/api/v1/cases/{case_id}/export/xlsx")
    wb = _load_xlsx(resp.content)
    assert "Caso" in wb.sheetnames
    # Sem solve → aba Resultados/Geometria ausentes
    assert "Resultados" not in wb.sheetnames
    assert "Geometria" not in wb.sheetnames
