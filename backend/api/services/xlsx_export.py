"""
Exportador Excel (.xlsx) com 3 abas + Diagnostics opcional (Fase 5 / Q6).

Estrutura:
  Aba "Caso":        metadados + inputs (segments + boundary + seabed)
  Aba "Resultados":  escalares + ProfileType + utilização + alert
  Aba "Geometria":   ≥ 5000 linhas (mesmo conteúdo do CSV)
  Aba "Diagnostics": (opcional) tabela de diagnostics estruturados se
                     houver (severity + confidence + code + título)

Estrutura da aba Diagnostics consistente com Memorial PDF (mesmo
ordenamento de colunas) — Q6 detail.

Usa openpyxl (já em requirements). Sem pandas para evitar overhead de
dependência transitiva.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.api.schemas.cases import CaseInput
from backend.api.services.case_hash import case_input_short_hash
from backend.solver.types import SolverResult


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_LABEL_FONT = Font(bold=True)


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_caso_sheet(wb: Workbook, case_input: CaseInput, hash_str: str) -> None:
    """Aba 1: Caso — metadados + inputs."""
    ws = wb.create_sheet("Caso")

    # Header / metadata
    ws["A1"] = "AncoPlat — Caso"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    rows = [
        ("Nome:", case_input.name),
        ("Descrição:", case_input.description or "—"),
        ("Hash do caso:", f"{hash_str}…"),
        ("Critério de utilização:", case_input.criteria_profile.value),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    # Boundary
    bc = case_input.boundary
    ws.cell(row=8, column=1, value="Boundary").font = Font(bold=True, size=12)
    boundary_rows = [
        ("Lâmina d'água sob âncora (m):", bc.h),
        ("Modo:", bc.mode.value),
        ("Input value:", bc.input_value),
        ("Profundidade do fairlead (m):", bc.startpoint_depth),
        ("Endpoint grounded:", bc.endpoint_grounded),
        ("Startpoint type:", bc.startpoint_type),
        ("Offset horz (m):", bc.startpoint_offset_horz),
        ("Offset vert (m):", bc.startpoint_offset_vert),
    ]
    for i, (label, value) in enumerate(boundary_rows, start=9):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    # Seabed
    seabed_row = 9 + len(boundary_rows) + 1
    ws.cell(row=seabed_row, column=1, value="Seabed").font = Font(bold=True, size=12)
    sb = case_input.seabed
    seabed_rows = [
        ("μ global:", sb.mu),
        ("Slope (rad):", sb.slope_rad),
        ("Slope (deg):", sb.slope_rad * 180 / 3.141592653589793),
    ]
    for i, (label, value) in enumerate(seabed_rows, start=seabed_row + 1):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    # Segmentos
    seg_row = seabed_row + len(seabed_rows) + 3
    ws.cell(row=seg_row, column=1, value=f"Segmentos ({len(case_input.segments)})").font = (
        Font(bold=True, size=12)
    )
    headers = [
        "#", "Line type", "Category", "Length (m)", "Diameter (m)",
        "w (N/m)", "EA (N)", "MBL (N)",
        "EA source", "μ override", "μ catálogo",
    ]
    _write_header_row(ws, seg_row + 1, headers)
    for i, seg in enumerate(case_input.segments, start=1):
        row = seg_row + 1 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=seg.line_type or "—")
        ws.cell(row=row, column=3, value=seg.category or "—")
        ws.cell(row=row, column=4, value=seg.length)
        ws.cell(row=row, column=5, value=seg.diameter)
        ws.cell(row=row, column=6, value=seg.w)
        ws.cell(row=row, column=7, value=seg.EA)
        ws.cell(row=row, column=8, value=seg.MBL)
        ws.cell(row=row, column=9, value=seg.ea_source)
        ws.cell(row=row, column=10, value=seg.mu_override)
        ws.cell(row=row, column=11, value=seg.seabed_friction_cf)

    _autosize_columns(ws, {1: 30, 2: 25, 3: 15, 4: 12, 5: 12, 6: 12,
                            7: 14, 8: 14, 9: 12, 10: 12, 11: 12})


def _build_resultados_sheet(wb: Workbook, result: SolverResult) -> None:
    """Aba 2: Resultados — escalares + ProfileType + utilização."""
    ws = wb.create_sheet("Resultados")

    ws["A1"] = "AncoPlat — Resultados"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    pt_value = (
        result.profile_type.value
        if result.profile_type is not None
        else "—"
    )
    rows = [
        ("Status:", result.status.value),
        ("Solver version:", result.solver_version),
        ("Profile Type:", pt_value),
        ("Utilization:", result.utilization),
        ("Alert level:", result.alert_level.value),
        ("", ""),
        ("Tração no fairlead (N):", result.fairlead_tension),
        ("Tração na âncora (N):", result.anchor_tension),
        ("Tração horizontal H (N):", result.H),
        ("X total (m):", result.total_horz_distance),
        ("L unstretched (m):", result.unstretched_length),
        ("L stretched (m):", result.stretched_length),
        ("L suspended (m):", result.total_suspended_length),
        ("L grounded (m):", result.total_grounded_length),
        ("Elongation (m):", result.elongation),
        ("", ""),
        ("Ângulo no fairlead vs horiz (rad):", result.angle_wrt_horz_fairlead),
        ("Ângulo no fairlead vs vert (rad):", result.angle_wrt_vert_fairlead),
        ("Ângulo na âncora vs horiz (rad):", result.angle_wrt_horz_anchor),
        ("Ângulo na âncora vs vert (rad):", result.angle_wrt_vert_anchor),
        ("Iterações usadas:", result.iterations_used),
        ("Anchor uplift severity:", result.anchor_uplift_severity),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=i, column=2, value=value)

    _autosize_columns(ws, {1: 35, 2: 20})


def _build_geometria_sheet(wb: Workbook, result: SolverResult) -> None:
    """Aba 3: Geometria — ≥ 5000 linhas com x, y, T_x, T_y, T_mag."""
    ws = wb.create_sheet("Geometria")

    headers = [
        "x (m)", "y (m)", "tension_x (N)", "tension_y (N)",
        "tension_magnitude (N)",
    ]
    _write_header_row(ws, 1, headers)

    for i in range(len(result.coords_x)):
        ws.cell(row=i + 2, column=1, value=result.coords_x[i])
        ws.cell(row=i + 2, column=2, value=result.coords_y[i])
        ws.cell(row=i + 2, column=3, value=result.tension_x[i])
        ws.cell(row=i + 2, column=4, value=result.tension_y[i])
        ws.cell(row=i + 2, column=5, value=result.tension_magnitude[i])

    _autosize_columns(ws, {1: 12, 2: 12, 3: 14, 4: 14, 5: 18})


def _build_diagnostics_sheet(wb: Workbook, result: SolverResult) -> None:
    """
    Aba 4 (condicional): Diagnostics — só se houver diagnostics
    estruturados. Estrutura consistente com Memorial PDF (Q6 detail):
    severity + confidence + code + title.
    """
    diagnostics = result.diagnostics or []
    if not diagnostics:
        return

    ws = wb.create_sheet("Diagnostics")

    headers = ["Code", "Severity", "Confidence", "Title", "Cause"]
    _write_header_row(ws, 1, headers)

    sev_colors = {
        "critical": "FEE2E2",
        "error": "FED7AA",
        "warning": "FEF3C7",
        "info": "DBEAFE",
    }

    for i, d in enumerate(diagnostics, start=1):
        row = i + 1
        ws.cell(row=row, column=1, value=d.get("code", "—"))
        sev_cell = ws.cell(row=row, column=2, value=d.get("severity", "—"))
        sev = d.get("severity", "")
        if sev in sev_colors:
            sev_cell.fill = PatternFill("solid", fgColor=sev_colors[sev])
        ws.cell(row=row, column=3, value=d.get("confidence", "—"))
        ws.cell(row=row, column=4, value=d.get("title", ""))
        ws.cell(row=row, column=5, value=d.get("cause", "")[:200])

    _autosize_columns(ws, {1: 32, 2: 12, 3: 12, 4: 50, 5: 60})


def build_xlsx(
    case_input: CaseInput, result: SolverResult | None,
) -> bytes:
    """
    Gera Excel com 3 abas mínimas + Diagnostics opcional (Q6 = a).

    AC do plano: 3 abas mínimas (Caso, Resultados, Geometria). Aba
    Diagnostics adicionada quando result.diagnostics não-vazio.

    Sem `result` (caso não-resolvido): só aba Caso. Resultados e
    Geometria ficam ausentes.
    """
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    if default is not None:
        wb.remove(default)

    hash_str = case_input_short_hash(case_input)
    _build_caso_sheet(wb, case_input, hash_str)

    if result is not None:
        _build_resultados_sheet(wb, result)
        _build_geometria_sheet(wb, result)
        _build_diagnostics_sheet(wb, result)  # condicional internamente

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build_xlsx"]
